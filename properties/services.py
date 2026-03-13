from datetime import datetime
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from .models import Property, Tenant


EXPECTED_COLUMNS = [
    'Tenant Name', 'Company Name', 'Property', 'Unit Number',
    'Rent Amount', 'Maintenance Charges', 'Parking Charges',
    'GST Number', 'Billing Cycle', 'Lease Start Date',
    'Lease End Date', 'Payment Due Day',
]


def import_tenants_from_excel(file):
    wb = load_workbook(file, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [{'row': 0, 'field': '', 'message': 'Empty file'}]

    headers = [str(h).strip() if h else '' for h in rows[0]]

    # Validate headers
    missing = [col for col in EXPECTED_COLUMNS if col not in headers]
    if missing:
        return [], [{'row': 1, 'field': '', 'message': f'Missing columns: {", ".join(missing)}'}]

    col_map = {h: i for i, h in enumerate(headers)}
    property_cache = {p.name.lower(): p for p in Property.objects.filter(is_active=True)}

    created = []
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        def get_val(col_name):
            idx = col_map.get(col_name)
            if idx is not None and idx < len(row):
                return row[idx]
            return None

        # Required fields
        tenant_name = str(get_val('Tenant Name') or '').strip()
        company_name = str(get_val('Company Name') or '').strip()
        property_name = str(get_val('Property') or '').strip()
        unit_number = str(get_val('Unit Number') or '').strip()

        if not all([tenant_name, company_name, property_name, unit_number]):
            errors.append({'row': row_num, 'field': 'required', 'message': 'Missing required fields (Tenant Name, Company Name, Property, Unit Number)'})
            continue

        # Property lookup
        prop = property_cache.get(property_name.lower())
        if not prop:
            errors.append({'row': row_num, 'field': 'Property', 'message': f'Property "{property_name}" not found'})
            continue

        # Amounts
        try:
            rent = Decimal(str(get_val('Rent Amount') or 0))
            if rent < 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            errors.append({'row': row_num, 'field': 'Rent Amount', 'message': 'Invalid rent amount'})
            continue

        try:
            maintenance = Decimal(str(get_val('Maintenance Charges') or 0))
        except (InvalidOperation, ValueError):
            maintenance = Decimal('0')

        try:
            parking = Decimal(str(get_val('Parking Charges') or 0))
        except (InvalidOperation, ValueError):
            parking = Decimal('0')

        # Billing cycle
        billing_cycle = str(get_val('Billing Cycle') or 'monthly').strip().lower()
        if billing_cycle not in ('monthly', 'quarterly'):
            errors.append({'row': row_num, 'field': 'Billing Cycle', 'message': f'Invalid billing cycle: {billing_cycle}'})
            continue

        # Dates
        lease_start = get_val('Lease Start Date')
        lease_end = get_val('Lease End Date')
        try:
            if isinstance(lease_start, datetime):
                lease_start = lease_start.date()
            else:
                lease_start = datetime.strptime(str(lease_start), '%Y-%m-%d').date()
            if isinstance(lease_end, datetime):
                lease_end = lease_end.date()
            else:
                lease_end = datetime.strptime(str(lease_end), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            errors.append({'row': row_num, 'field': 'Dates', 'message': 'Invalid date format (use YYYY-MM-DD)'})
            continue

        # Payment due day
        try:
            due_day = int(get_val('Payment Due Day') or 5)
            if not (1 <= due_day <= 28):
                due_day = 5
        except (ValueError, TypeError):
            due_day = 5

        gst_number = str(get_val('GST Number') or '').strip()

        # Duplicate check
        if Tenant.objects.filter(building=prop, unit_number=unit_number).exists():
            errors.append({'row': row_num, 'field': 'Unit Number', 'message': f'Tenant already exists for {prop.name} - {unit_number}'})
            continue

        tenant = Tenant.objects.create(
            building=prop,
            name=tenant_name,
            company_name=company_name,
            unit_number=unit_number,
            contact_email='',
            contact_phone='',
            gst_number=gst_number,
            rent_amount=rent,
            maintenance_charges=maintenance,
            parking_charges=parking,
            billing_cycle=billing_cycle,
            lease_start_date=lease_start,
            lease_end_date=lease_end,
            payment_due_day=due_day,
        )
        created.append(tenant)

    wb.close()
    return created, errors
